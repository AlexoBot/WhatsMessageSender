from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from datetime import datetime
import time, os, getpass
from dotenv import load_dotenv


#  ///////////////// WhatsApp Message Scheduler v1.0 //////////////////////////
# ///////////////////          For Chrome          ///////////////////////////



def WaitTimer(total_seconds):
    update_interval = 30  # Update the message every 30 seconds
    while total_seconds > 0:
        if total_seconds >= update_interval:
            print(f'{total_seconds} seconds remaining...')
            time.sleep(update_interval)
            total_seconds -= update_interval
        else:
            print(f'{total_seconds} seconds remaining...')
            time.sleep(total_seconds)
            total_seconds = 0
    print('Task Done\n')

def GeneratePlanning(template_path,File_path):
    print('== Generating Planning ==')
    wb = load_workbook(template_path)
    ws = wb.active
    current_date = datetime.now()
    ws['B2'] = current_date.strftime('%A %d %B %Y')
    ws['B12'] = 'Pending or suggestions'
    ws['B16'] = 'Follow ups'
    wb.save(File_path)
    print(f'File generated! saved as {File_path}')
    print()

def driverConfigure(Chrome_user_dir_path,chrome_for_testing_path,chrome_driver_path,profile_name):
    
    print("Logging in and loading WhatsApp Web...")
    
    options = webdriver.ChromeOptions()
    options.binary_location = chrome_for_testing_path
    options.add_argument(Chrome_user_dir_path)
    options.add_argument(f"--profile-directory={profile_name}")
    service = Service(chrome_driver_path)

    driver = webdriver.Chrome(service=service, options=options)

    driver.get('https://web.whatsapp.com')
    print("WhatsApp Web loaded successfully.")
    time.sleep(25)
    print()
    return driver

def SearchGroupOrContact(driver,Search_box_xpath,Group_or_chat_name):
    print("== Searching group or contact ==")
    search_box = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, Search_box_xpath))
    )

    # Scroll into view
    driver.execute_script("arguments[0].scrollIntoView(true);", search_box)
    time.sleep(0.5)

    # Focus and clear
    driver.execute_script("arguments[0].focus();", search_box)
    driver.execute_script("arguments[0].value = '';", search_box)
    time.sleep(0.3)

    # Set value and trigger input events using JavaScript
    driver.execute_script(f"""
        var element = arguments[0];
        element.value = '{Group_or_chat_name}';
        element.dispatchEvent(new Event('input', {{ bubbles: true }}));
        element.dispatchEvent(new Event('change', {{ bubbles: true }}));
    """, search_box)

    time.sleep(1)

    # Press Enter using JavaScript
    driver.execute_script("""
        var event = new KeyboardEvent('keydown', {{
            key: 'Enter',
            code: 'Enter',
            keyCode: 13,
            bubbles: true,
            cancelable: true
        }});
        arguments[0].dispatchEvent(event);

        var event2 = new KeyboardEvent('keyup', {{
            key: 'Enter',
            code: 'Enter',
            keyCode: 13,
            bubbles: true,
            cancelable: true
        }});
        arguments[0].dispatchEvent(event2);
    """, search_box)

    time.sleep(5)
    print()

def SendMessage(driver, Message_box_xpath,Message_to_send):
    print("== Sending message ==")
    message_box = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, Message_box_xpath))
    )

    # Scroll into view
    driver.execute_script("arguments[0].scrollIntoView(true);", message_box)
    time.sleep(0.5)

    # Focus and clear
    driver.execute_script("arguments[0].focus();", message_box)
    driver.execute_script("arguments[0].value = '';", message_box)
    time.sleep(0.3)

    # Set value and trigger input events using JavaScript
    driver.execute_script(f"""
        var element = arguments[0];
        element.value = '{Message_to_send}';
        element.dispatchEvent(new Event('input', {{ bubbles: true }}));
        element.dispatchEvent(new Event('change', {{ bubbles: true }}));
    """, message_box)

    time.sleep(1)

    # Press Enter using JavaScript
    driver.execute_script("""
        var event = new KeyboardEvent('keydown', {{
            key: 'Enter',
            code: 'Enter',
            keyCode: 13,
            bubbles: true,
            cancelable: true
        }});
        arguments[0].dispatchEvent(event);

        var event2 = new KeyboardEvent('keyup', {{
            key: 'Enter',
            code: 'Enter',
            keyCode: 13,
            bubbles: true,
            cancelable: true
        }});
        arguments[0].dispatchEvent(event2);
    """, message_box)

    print('Message sent')
    print()

def SendAttachment(driver, Attach_button_xpath, Document_button_xpath, File_path, Attached_send_button):
    print('== Sending a document ==')

    try:
        absolute_file_path = os.path.abspath(File_path)
        print(f'Attempting to attach file: {absolute_file_path}')

        if not os.path.exists(absolute_file_path):
            raise FileNotFoundError(f'File not found: {absolute_file_path}')

        # 1) Click the paperclip
        print('Clicking attach button...')
        attach_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, Attach_button_xpath))
        )
        attach_button.click()
        time.sleep(1)

        # 2) Click the "Document" icon (this is where your Document_button_xpath should point)
        print('Clicking document button...')
        document_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, Document_button_xpath))
        )
        document_button.click()

        # 3) Find the correct file input for documents
        # WhatsApp usually has several inputs; we pick one that accepts documents (xls/xlsx)
        print('Locating document file input...')
        file_inputs = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'input[type="file"]'))
        )

        doc_input = None
        for inp in file_inputs:
            accept = (inp.get_attribute("accept") or "").lower()
            print(f"Found file input with accept='{accept}'")
            if (
                ".xls" in accept
                or ".xlsx" in accept
                or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in accept
                or accept.strip() == ""  # some builds use empty accept for "any file"
            ):
                doc_input = inp
                break

        # Fallback: last input if none clearly matches
        if doc_input is None:
            print("No specific document input found, using last input as fallback.")
            doc_input = file_inputs[-1]

        # 4) Send file path
        print('Sending file path to document input...')
        doc_input.send_keys(absolute_file_path)

        # 5) Wait for preview to appear (the attach preview popup)
        #   You can fine-tune this selector; for now we just wait a bit
        print('Waiting for WhatsApp to process file and show preview...')
        time.sleep(5)

        # 6) Click the Send button
        print('Clicking send button...')
        send_button = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, Attached_send_button))
        )

        # Try normal click first
        try:
            send_button.click()
        except Exception:
            print('Regular click failed, trying JavaScript click...')
            driver.execute_script("arguments[0].click();", send_button)

        print('Document sent successfully!')
        time.sleep(3)

    except FileNotFoundError as e:
        print(f'ERROR: {e}')
        raise
    except Exception as e:
        print(f'ERROR in SendAttachment: {type(e).__name__}: {e}')
        print('Current URL:', driver.current_url)
        try:
            screenshot_path = os.path.join(os.path.dirname(File_path), 'error_screenshot.png')
            driver.save_screenshot(screenshot_path)
            print(f'Screenshot saved to: {screenshot_path}')
        except:
            pass
        raise

    print()


if __name__ == '__main__':
    #=====================================================================================================================================
    #PERSONAL VARIABLES:
    load_dotenv(".env")
    username = getpass.getuser()
    target_file_path = os.getenv("ONEDRIVEPATH").format(username=username)
    #====================================================================================================================================
    #PATH VARIABLES:
    load_dotenv(".env.defaults")
    profile_name = os.getenv("PROFILENAME")
    chrome_for_testing_path = os.getenv("CHROMETESTINGPATH")
    chrome_driver_path = os.getenv("CHROMEDRIVERPATH")
    Chrome_user_dir_path = os.getenv("CHROMEUSERDIRPATH").format(username=username, profile_name=profile_name)
    #====================================================================================================================================
    #NAMES, DATES & TEXT VARIABLES:
    current_date = datetime.now()
    formatted_date_short = current_date.strftime('%b-%d-%y')
    Group_or_chat_name = 'El grupo de las ardillas'
    Message_to_send = "TESTING, MESSAGES WORK"
    #====================================================================================================================================
    #FILES AND DOCS VARIABLES:
    template_name = 'DailyPlanning template.xlsx'
    new_filename = f"DailyPlanning {formatted_date_short}.xlsx"
    template_path = os.path.join(target_file_path, template_name)
    File_path = os.path.join(target_file_path, new_filename)
    #====================================================================================================================================
    #XPATHS VARIABLES:
    Search_box_xpath = os.getenv("SEARCHBOX")
    Message_box_xpath = os.getenv("MESSAGEBOX")
    Attach_button_xpath =os.getenv("ATTACHBUTTON")
    Document_button_xpath = os.getenv("DOCUMENTBUTTON")
    Attached_send_button = os.getenv("SENDATTACHMENT")

    print('\n','+'*10,' Running Tests With: ','+'*10,'\n')
    print(f'target_file_path:      {target_file_path}')
    print(f'Group_or_chat_name:    {Group_or_chat_name}')
    print(f'Message_to_send:       {Message_to_send}')
    print(f'template_name:         {template_name}\n')
    print('+'*42)

    GeneratePlanning(template_path,File_path)
    WaitTimer(1)
    driver = driverConfigure(Chrome_user_dir_path,chrome_for_testing_path,chrome_driver_path,profile_name)
    SearchGroupOrContact(driver, Search_box_xpath,Group_or_chat_name)
    SendMessage(driver, Message_box_xpath,Message_to_send)
    SendAttachment(driver, Attach_button_xpath,Document_button_xpath,File_path,Attached_send_button)
    WaitTimer(30)

    driver.quit()
else:
    print('Functions Imported Correctly')